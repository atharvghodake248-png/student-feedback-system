
from textblob import TextBlob
import re
from django.shortcuts import get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.contrib.auth import authenticate, login, logout
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.db import transaction
from django.db.models import Avg, Count, Q
import json
from datetime import datetime
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .models import (
    CustomUser, Student, Teacher, Subject, Branch, Year,
    Semester, Feedback, FeedbackSummary, Division, TeacherSubject
)

# ==================== SENTIMENT ANALYSIS ====================

def analyze_sentiment(text):
    """Analyze sentiment of text and return sentiment label and score"""
    if not text or text.strip() == '':
        return None, 0.0
    
    text_lower = text.lower()
    
    negative_keywords = [
        'not', 'no', 'bad', 'poor', 'worst', 'terrible', 'awful',
        'useless', 'waste', 'boring', 'confusing', 'difficult',
        'never', 'late', 'absent', 'rude', 'unprofessional'
    ]
    
    has_negative_keywords = any(keyword in text_lower for keyword in negative_keywords)
    
    blob = TextBlob(text)
    polarity = blob.sentiment.polarity
    
    if has_negative_keywords and polarity <= 0:
        return 'negative', polarity
    elif polarity > 0.05:
        return 'positive', polarity
    elif polarity < -0.05:
        return 'negative', polarity
    else:
        return 'neutral', polarity

# ==================== AUTHENTICATION ====================

@csrf_exempt
@require_http_methods(["POST"])
def login_view(request):
    """User login endpoint - SUPPORTS Username, PRN, Employee ID"""
    try:
        data = json.loads(request.body)
        login_id = data.get('username', '').strip()
        password = data.get('password', '').strip()
        
        if not login_id or not password:
            return JsonResponse({'error': 'Username and password are required'}, status=400)
        
        user = None
        actual_username = None
        
        try:
            user = CustomUser.objects.get(username=login_id)
            actual_username = user.username
        except CustomUser.DoesNotExist:
            try:
                user = CustomUser.objects.get(prn_number=login_id)
                actual_username = user.username
            except CustomUser.DoesNotExist:
                try:
                    student = Student.objects.get(prn_number=login_id)
                    user = student.user
                    actual_username = user.username
                except Student.DoesNotExist:
                    try:
                        teacher = Teacher.objects.get(employee_id=login_id)
                        user = teacher.user
                        actual_username = user.username
                    except Teacher.DoesNotExist:
                        pass
        
        if not user or not actual_username:
            return JsonResponse({'error': 'Invalid credentials'}, status=401)
        
        authenticated_user = authenticate(request, username=actual_username, password=password)
        
        if authenticated_user:
            login(request, authenticated_user)
            
            is_class_teacher = False
            if authenticated_user.user_type == 'teacher':
                try:
                    teacher_profile = authenticated_user.teacher_profile
                    is_class_teacher = teacher_profile.is_class_teacher
                except:
                    pass
            
            return JsonResponse({
                'success': True,
                'user_type': authenticated_user.user_type,
                'user_id': authenticated_user.id,
                'username': authenticated_user.username,
                'full_name': authenticated_user.get_full_name(),
                'is_class_teacher': is_class_teacher
            })
        
        return JsonResponse({'error': 'Invalid credentials'}, status=401)
        
    except Exception as e:
        import traceback
        print("LOGIN ERROR:", traceback.format_exc())
        return JsonResponse({'error': 'Login failed. Please try again.'}, status=500)

@csrf_exempt
def logout_view(request):
    """User logout endpoint"""
    logout(request)
    return JsonResponse({'success': True, 'message': 'Logged out successfully'})

# ==================== STUDENT VIEWS ====================

def student_dashboard(request):
    """Get student dashboard data"""
    try:
        username = request.GET.get('username')
        if not username:
            return JsonResponse({'error': 'Username required'}, status=400)
        
        user = CustomUser.objects.get(username=username, user_type='student')
        student = user.student_profile
        
        # Refresh student data from database
        student.refresh_from_db()
        
        feedback_count = Feedback.objects.filter(student=student).count()
        
        recent_feedback = Feedback.objects.filter(student=student).select_related(
            'subject', 'teacher', 'teacher__user'
        ).order_by('-created_at')[:5]
        
        recent_data = [{
            'subject': fb.subject.name,
            'teacher': fb.teacher.user.get_full_name(),
            'rating': fb.overall_satisfaction,
            'date': fb.created_at.strftime('%Y-%m-%d')
        } for fb in recent_feedback]
        
        division_str = f"Division {student.division.name}" if student.division else "N/A"
        
        return JsonResponse({
            'success': True,
            'student': {
                'prn': student.prn_number,
                'name': user.get_full_name(),
                'year': student.year.name,
                'branch': student.branch.name,
                'semester': f"Semester {student.semester.number}",
                'division': division_str,
                'email': user.email
            },
            'feedback_submitted': feedback_count,
            'recent_feedback': recent_data
        })
        
    except CustomUser.DoesNotExist:
        return JsonResponse({'error': 'Student not found'}, status=404)
    except Exception as e:
        import traceback
        print("STUDENT DASHBOARD ERROR:", traceback.format_exc())
        return JsonResponse({'error': str(e)}, status=500)

def get_subjects(request, year_id, branch_id, semester_id):
    """Get subjects for specific year, branch, and semester"""
    try:
        subjects = Subject.objects.filter(
            semester__year_id=year_id,
            branch_id=branch_id,
            semester_id=semester_id
        ).select_related('division').order_by('code')
        
        return JsonResponse({
            'success': True,
            'subjects': [
                {
                    'id': s.id,
                    'code': s.code,
                    'name': s.name,
                    'credits': s.credits,
                    'division': s.division.name if s.division else 'Common'
                }
                for s in subjects
            ]
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def get_all_subjects(request):
    """Get all subjects with full details for admin view - COMPLETE FIX"""
    try:
        subjects = Subject.objects.all().select_related(
            'semester', 'semester__year', 'branch', 'division'
        ).order_by('semester__year__name', 'branch__name', 'semester__number', 'code')
        
        subjects_list = []
        for s in subjects:
            subjects_list.append({
                'id': s.id,
                'code': s.code,
                'name': s.name,
                'credits': s.credits,
                'year_id': s.semester.year.id,
                'year': s.semester.year.name,  # For display
                'branch_id': s.branch.id,
                'branch': s.branch.name,  # For display
                'semester_id': s.semester.id,
                'semester': s.semester.number,  # For display
                'division_id': s.division.id if s.division else None,
                'division': s.division.name if s.division else 'Common',
            })
        
        return JsonResponse({
            'success': True,
            'subjects': subjects_list
        })
    except Exception as e:
        import traceback
        print("GET ALL SUBJECTS ERROR:", traceback.format_exc())
        return JsonResponse({'error': str(e)}, status=500)

def get_all_students(request):
    """Get all students for admin view WITH DIVISION"""
    try:
        students = Student.objects.all().select_related(
            'user', 'year', 'branch', 'semester', 'division'
        ).order_by('prn_number')
        
        return JsonResponse({
            'success': True,
            'students': [{
                'id': s.id,
                'prn_number': s.prn_number,
                'name': s.user.get_full_name(),
                'username': s.user.username,
                'email': s.user.email,
                'year': s.year.name,
                'year_id': s.year.id,
                'branch': s.branch.name,
                'branch_id': s.branch.id,
                'semester': f"Semester {s.semester.number}",
                'semester_id': s.semester.id,
                'division': s.division.name if s.division else 'N/A',
                'division_id': s.division.id if s.division else None,
                'is_active': s.user.is_active
            } for s in students]
        })
    except Exception as e:
        import traceback
        print("GET ALL STUDENTS ERROR:", traceback.format_exc())
        return JsonResponse({'error': str(e)}, status=500)

def get_all_teachers(request):
    """Get all teachers for admin view"""
    try:
        teachers = Teacher.objects.all().select_related(
            'user', 'department'
        ).prefetch_related('subjects').order_by('employee_id')
        
        return JsonResponse({
            'success': True,
            'teachers': [{
                'id': t.id,
                'employee_id': t.employee_id,
                'name': t.user.get_full_name(),
                'username': t.user.username,
                'email': t.user.email,
                'department': t.department.name if t.department else 'N/A',
                'department_id': t.department.id if t.department else None,
                'subjects_count': t.subjects.count(),
                'is_active': t.user.is_active,
                'is_class_teacher': t.is_class_teacher
            } for t in teachers]
        })
    except Exception as e:
        import traceback
        print("GET ALL TEACHERS ERROR:", traceback.format_exc())
        return JsonResponse({'error': str(e)}, status=500)

def manage_access(request):
    """Get all students and teachers for admin management"""
    try:
        students = Student.objects.all().select_related(
            'user', 'year', 'branch', 'semester', 'division'
        )
        teachers = Teacher.objects.all().select_related(
            'user', 'department'
        ).prefetch_related('subjects')
        
        students_data = [{
            'id': s.id,
            'prn': s.prn_number,
            'name': s.user.get_full_name(),
            'username': s.user.username,
            'email': s.user.email,
            'year': s.year.name,
            'branch': s.branch.name,
            'semester': f"Semester {s.semester.number}",
            'division': s.division.name if s.division else 'N/A',
            'is_active': s.user.is_active
        } for s in students]
        
        teachers_data = [{
            'id': t.id,
            'employee_id': t.employee_id,
            'name': t.user.get_full_name(),
            'username': t.user.username,
            'email': t.user.email,
            'department': t.department.name if t.department else 'N/A',
            'subjects_count': t.subjects.count(),
            'is_active': t.user.is_active,
            'is_class_teacher': t.is_class_teacher
        } for t in teachers]
        
        return JsonResponse({
            'success': True,
            'students': students_data,
            'teachers': teachers_data
        })
        
    except Exception as e:
        import traceback
        print("MANAGE ACCESS ERROR:", traceback.format_exc())
        return JsonResponse({'error': str(e)}, status=500)

# ==================== DELETE OPERATIONS ====================

@csrf_exempt
@require_http_methods(["DELETE"])
def delete_student(request, student_id):
    """Delete a student (soft delete recommended)"""
    try:
        student = get_object_or_404(Student, id=student_id)
        
        student.user.is_active = False
        student.user.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Student {student.prn_number} deactivated successfully'
        })
        
    except Exception as e:
        import traceback
        print("DELETE STUDENT ERROR:", traceback.format_exc())
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@require_http_methods(["DELETE"])
def delete_teacher(request, teacher_id):
    """Delete a teacher (soft delete recommended)"""
    try:
        teacher = get_object_or_404(Teacher, id=teacher_id)
        
        teacher.user.is_active = False
        teacher.user.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Teacher {teacher.employee_id} deactivated successfully'
        })
        
    except Exception as e:
        import traceback
        print("DELETE TEACHER ERROR:", traceback.format_exc())
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@require_http_methods(["DELETE"])
def delete_subject(request, subject_id):
    """Delete a subject"""
    try:
        subject = get_object_or_404(Subject, id=subject_id)
        
        feedback_count = Feedback.objects.filter(subject=subject).count()
        
        if feedback_count > 0:
            return JsonResponse({
                'error': f'Cannot delete subject. {feedback_count} feedback entries exist for this subject.'
            }, status=400)
        
        subject_code = subject.code
        subject.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Subject {subject_code} deleted successfully'
        })
        
    except Exception as e:
        import traceback
        print("DELETE SUBJECT ERROR:", traceback.format_exc())
        return JsonResponse({'error': str(e)}, status=500)

# ==================== BULK OPERATIONS ====================

@csrf_exempt
@require_http_methods(["POST"])
def bulk_add_students(request):
    """Bulk add students from CSV/JSON data"""
    try:
        data = json.loads(request.body)
        students_data = data.get('students', [])
        
        if not students_data:
            return JsonResponse({'error': 'No student data provided'}, status=400)
        
        added_students = []
        errors = []
        
        for idx, student_data in enumerate(students_data):
            try:
                required = ['prn_number', 'email', 'first_name', 'last_name', 
                           'password', 'year_id', 'branch_id', 'semester_id', 'division_id']
                
                missing = [f for f in required if f not in student_data]
                if missing:
                    errors.append({
                        'row': idx + 1,
                        'prn': student_data.get('prn_number', 'Unknown'),
                        'error': f'Missing fields: {", ".join(missing)}'
                    })
                    continue
                
                if CustomUser.objects.filter(prn_number=student_data['prn_number']).exists():
                    errors.append({
                        'row': idx + 1,
                        'prn': student_data['prn_number'],
                        'error': 'PRN already exists'
                    })
                    continue
                
                with transaction.atomic():
                    username = student_data.get('username', student_data['prn_number'])
                    
                    if CustomUser.objects.filter(username=username).exists():
                        username = student_data['prn_number']
                    
                    user = CustomUser.objects.create_user(
                        username=username,
                        email=student_data['email'],
                        first_name=student_data['first_name'],
                        last_name=student_data['last_name'],
                        password=student_data['password'],
                        user_type='student',
                        prn_number=student_data['prn_number']
                    )
                    
                    Student.objects.create(
                        user=user,
                        prn_number=student_data['prn_number'],
                        year_id=student_data['year_id'],
                        branch_id=student_data['branch_id'],
                        semester_id=student_data['semester_id'],
                        division_id=student_data['division_id']
                    )
                    
                    added_students.append(student_data['prn_number'])
                    
            except Exception as e:
                errors.append({
                    'row': idx + 1,
                    'prn': student_data.get('prn_number', 'Unknown'),
                    'error': str(e)
                })
        
        return JsonResponse({
            'success': True,
            'message': f'Bulk add completed. {len(added_students)} students added, {len(errors)} errors.',
            'added_count': len(added_students),
            'added_students': added_students,
            'error_count': len(errors),
            'errors': errors
        })
        
    except Exception as e:
        import traceback
        print("BULK ADD STUDENTS ERROR:", traceback.format_exc())
        return JsonResponse({'error': str(e)}, status=500)

# ==================== STATISTICS & REPORTS ====================

def get_admin_statistics(request):
    """Get overall system statistics for admin dashboard"""
    try:
        total_students = Student.objects.count()
        total_teachers = Teacher.objects.count()
        total_subjects = Subject.objects.count()
        total_feedback = Feedback.objects.count()
        
        active_students = Student.objects.filter(user__is_active=True).count()
        active_teachers = Teacher.objects.filter(user__is_active=True).count()
        
        avg_ratings = Feedback.objects.aggregate(
            avg_overall=Avg('overall_satisfaction'),
            avg_teaching=Avg('teaching_effectiveness'),
            avg_content=Avg('course_content')
        )
        
        branch_distribution = Student.objects.values('branch__name').annotate(
            count=Count('id')
        ).order_by('-count')
        
        year_distribution = Student.objects.values('year__name').annotate(
            count=Count('id')
        ).order_by('year__name')
        
        recent_feedback = Feedback.objects.select_related(
            'student', 'student__user', 'teacher', 'teacher__user', 'subject'
        ).order_by('-created_at')[:10]
        
        recent_feedback_data = [{
            'id': fb.id,
            'student': fb.student.user.get_full_name() if not fb.is_anonymous else 'Anonymous',
            'teacher': fb.teacher.user.get_full_name(),
            'subject': fb.subject.name,
            'rating': fb.overall_satisfaction,
            'date': fb.created_at.strftime('%Y-%m-%d %H:%M')
        } for fb in recent_feedback]
        
        return JsonResponse({
            'success': True,
            'statistics': {
                'total_students': total_students,
                'total_teachers': total_teachers,
                'total_subjects': total_subjects,
                'total_feedback': total_feedback,
                'active_students': active_students,
                'active_teachers': active_teachers,
                'average_ratings': {
                    'overall': round(avg_ratings['avg_overall'] or 0, 2),
                    'teaching': round(avg_ratings['avg_teaching'] or 0, 2),
                    'content': round(avg_ratings['avg_content'] or 0, 2)
                },
                'branch_distribution': list(branch_distribution),
                'year_distribution': list(year_distribution)
            },
            'recent_feedback': recent_feedback_data
        })
        
    except Exception as e:
        import traceback
        print("GET ADMIN STATISTICS ERROR:", traceback.format_exc())
        return JsonResponse({'error': str(e)}, status=500)

def download_all_feedback_report(request):
    """Download comprehensive feedback report for all teachers"""
    try:
        feedbacks = Feedback.objects.all().select_related(
            'student', 'student__user', 'student__branch', 'student__year', 'student__division',
            'teacher', 'teacher__user', 'subject', 'semester', 'semester__year'
        ).order_by('-created_at')
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="all_feedback_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        response.write('\ufeff')
        
        writer = csv.writer(response)
        writer.writerow([
            'Feedback ID', 'Date', 'Student PRN', 'Student Name', 'Year', 'Branch', 
            'Division', 'Teacher ID', 'Teacher Name', 'Subject Code', 'Subject Name', 
            'Semester', 'Teaching', 'Content', 'Interaction', 'Assignment', 'Overall',
            'Comments', 'Comment Sentiment', 'Suggestions', 'Suggestion Sentiment', 'Anonymous'
        ])
        
        for fb in feedbacks:
            writer.writerow([
                fb.id,
                fb.created_at.strftime('%d-%m-%Y %H:%M'),
                fb.student.prn_number if not fb.is_anonymous else 'Anonymous',
                fb.student.user.get_full_name() if not fb.is_anonymous else 'Anonymous',
                fb.student.year.name if not fb.is_anonymous else 'N/A',
                fb.student.branch.name if not fb.is_anonymous else 'N/A',
                fb.student.division.name if (not fb.is_anonymous and fb.student.division) else 'N/A',
                fb.teacher.employee_id,
                fb.teacher.user.get_full_name(),
                fb.subject.code,
                fb.subject.name,
                f"{fb.semester.year.name} Sem-{fb.semester.number}",
                fb.teaching_effectiveness,
                fb.course_content,
                fb.interaction_quality,
                fb.assignment_feedback,
                fb.overall_satisfaction,
                fb.comments or 'No comments',
                fb.comment_sentiment or 'Not analyzed',
                fb.suggestions or 'No suggestions',
                fb.suggestion_sentiment or 'Not analyzed',
                'Yes' if fb.is_anonymous else 'No'
            ])
        
        return response
        
    except Exception as e:
        import traceback
        print("DOWNLOAD ALL FEEDBACK ERROR:", traceback.format_exc())
        return JsonResponse({'error': str(e)}, status=500)

# ==================== SEARCH & FILTER ====================

def search_users(request):
    """Search for students and teachers"""
    try:
        query = request.GET.get('q', '').strip()
        user_type = request.GET.get('type', 'all')
        
        if not query:
            return JsonResponse({'error': 'Search query required'}, status=400)
        
        results = {
            'students': [],
            'teachers': []
        }
        
        if user_type in ['student', 'all']:
            students = Student.objects.filter(
                Q(prn_number__icontains=query) |
                Q(user__first_name__icontains=query) |
                Q(user__last_name__icontains=query) |
                Q(user__email__icontains=query)
            ).select_related('user', 'year', 'branch', 'semester', 'division')[:20]
            
            results['students'] = [{
                'id': s.id,
                'prn': s.prn_number,
                'name': s.user.get_full_name(),
                'email': s.user.email,
                'year': s.year.name,
                'branch': s.branch.name,
                'division': s.division.name if s.division else 'N/A'
            } for s in students]
        
        if user_type in ['teacher', 'all']:
            teachers = Teacher.objects.filter(
                Q(employee_id__icontains=query) |
                Q(user__first_name__icontains=query) |
                Q(user__last_name__icontains=query) |
                Q(user__email__icontains=query)
            ).select_related('user', 'department')[:20]
            
            results['teachers'] = [{
                'id': t.id,
                'employee_id': t.employee_id,
                'name': t.user.get_full_name(),
                'email': t.user.email,
                'department': t.department.name if t.department else 'N/A'
            } for t in teachers]
        
        return JsonResponse({
            'success': True,
            'query': query,
            'results': results,
            'total_found': len(results['students']) + len(results['teachers'])
        })
        
    except Exception as e:
        import traceback
        print("SEARCH USERS ERROR:", traceback.format_exc())
        return JsonResponse({'error': str(e)}, status=500)

# ==================== TEACHER-SUBJECT ASSIGNMENT ====================

def get_subjects_by_class(request):
    """Get subjects for specific year, branch, semester, division"""
    try:
        year_id = request.GET.get('year_id')
        branch_id = request.GET.get('branch_id')
        semester_id = request.GET.get('semester_id')
        division_id = request.GET.get('division_id')
        
        if not all([year_id, branch_id, semester_id]):
            return JsonResponse({
                'success': False,
                'error': 'Year, Branch, and Semester are required'
            }, status=400)
        
        subjects = Subject.objects.filter(
            semester__year_id=year_id,
            branch_id=branch_id,
            semester_id=semester_id
        )
        
        if division_id:
            subjects = subjects.filter(
                Q(division_id=division_id) | Q(division_id__isnull=True)
            )
        else:
            subjects = subjects.filter(division_id__isnull=True)
        
        subjects_list = [{
            'id': s.id,
            'code': s.code,
            'name': s.name,
            'credits': s.credits,
            'division': s.division.name if s.division else None
        } for s in subjects]
        
        return JsonResponse({
            'success': True,
            'subjects': subjects_list
        })
    except Exception as e:
        import traceback
        print("GET SUBJECTS BY CLASS ERROR:", traceback.format_exc())
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@csrf_exempt
@require_http_methods(["POST"])
def assign_subject_to_teacher(request):
    """Assign a subject to a teacher"""
    try:
        data = json.loads(request.body)
        teacher_id = data.get('teacher_id')
        subject_id = data.get('subject_id')
        
        if not teacher_id or not subject_id:
            return JsonResponse({
                'success': False,
                'error': 'Teacher and Subject are required'
            }, status=400)
        
        try:
            teacher = Teacher.objects.get(id=teacher_id)
        except Teacher.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Teacher not found'
            }, status=404)
        
        try:
            subject = Subject.objects.get(id=subject_id)
        except Subject.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Subject not found'
            }, status=404)
        
        if TeacherSubject.objects.filter(teacher_id=teacher_id, subject_id=subject_id).exists():
            return JsonResponse({
                'success': False,
                'error': 'This subject is already assigned to this teacher'
            }, status=400)
        
        TeacherSubject.objects.create(teacher_id=teacher_id, subject_id=subject_id)
        
        return JsonResponse({
            'success': True,
            'message': f'Subject {subject.code} assigned to {teacher.user.get_full_name()} successfully'
        })
    except Exception as e:
        import traceback
        print("ASSIGN SUBJECT TO TEACHER ERROR:", traceback.format_exc())
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

def get_teacher_subjects(request, teacher_id):
    """Get all subjects assigned to a specific teacher"""
    try:
        try:
            teacher = Teacher.objects.get(id=teacher_id)
        except Teacher.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Teacher not found'
            }, status=404)
        
        assignments = TeacherSubject.objects.filter(
            teacher_id=teacher_id
        ).select_related(
            'subject',
            'subject__semester',
            'subject__semester__year',
            'subject__branch',
            'subject__division'
        ).order_by('-assigned_date')
        
        assignments_list = [{
            'id': ts.id,
            'subject_id': ts.subject.id,
            'subject_code': ts.subject.code,
            'subject_name': ts.subject.name,
            'credits': ts.subject.credits,
            'year': ts.subject.semester.year.name,
            'branch': ts.subject.branch.name,
            'semester': ts.subject.semester.number,
            'division': ts.subject.division.name if ts.subject.division else None,
            'assigned_date': ts.assigned_date.strftime('%Y-%m-%d')
        } for ts in assignments]
        
        return JsonResponse({
            'success': True,
            'assignments': assignments_list
        })
    except Exception as e:
        import traceback
        print("GET TEACHER SUBJECTS ERROR:", traceback.format_exc())
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@csrf_exempt
@require_http_methods(["DELETE"])
def remove_subject_assignment(request, assignment_id):
    """Remove a teacher-subject assignment"""
    try:
        try:
            assignment = TeacherSubject.objects.get(id=assignment_id)
        except TeacherSubject.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Assignment not found'
            }, status=404)
        
        subject_code = assignment.subject.code
        teacher_name = assignment.teacher.user.get_full_name()
        
        assignment.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Assignment removed: {subject_code} from {teacher_name}'
        })
    except Exception as e:
        import traceback
        print("REMOVE SUBJECT ASSIGNMENT ERROR:", traceback.format_exc())
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

def get_all_teacher_subject_assignments(request):
    """Get all teacher-subject assignments - FIXED TO MATCH FRONTEND"""
    try:
        assignments = TeacherSubject.objects.all().select_related(
            'teacher',
            'teacher__user',
            'subject',
            'subject__semester',
            'subject__semester__year',
            'subject__branch',
            'subject__division'
        ).order_by('teacher__employee_id', 'subject__code')
        
        assignments_list = []
        for ts in assignments:
            # Get teacher's full name
            teacher_full_name = f"{ts.teacher.user.first_name} {ts.teacher.user.last_name}".strip()
            if not teacher_full_name:
                teacher_full_name = ts.teacher.user.username
            
            # THIS MATCHES YOUR FRONTEND EXACTLY!
            assignments_list.append({
                'id': ts.id,
                'teacher_id': ts.teacher.id,
                'teacher_name': teacher_full_name,  # Frontend uses this
                'teacher_employee_id': ts.teacher.employee_id,  # THIS WAS MISSING!
                'subject_id': ts.subject.id,
                'subject_name': ts.subject.name,  # Frontend uses this
                'subject_code': ts.subject.code,  # Frontend uses this
                'code': ts.subject.code,
                'credits': ts.subject.credits,
                'year': ts.subject.semester.year.name,  # Frontend uses this
                'branch': ts.subject.branch.name,  # Frontend uses this
                'semester': ts.subject.semester.number,  # Frontend uses this
                'division': ts.subject.division.name if ts.subject.division else 'Common',
                'assigned_date': ts.assigned_date.strftime('%Y-%m-%d')
            })
        
        return JsonResponse({
            'success': True,
            'assignments': assignments_list,
            'total': len(assignments_list)
        })
    except Exception as e:
        import traceback
        print("GET ALL TEACHER SUBJECT ASSIGNMENTS ERROR:", traceback.format_exc())
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

# ==================== HEALTH CHECK ====================

def health_check(request):
    """API health check endpoint"""
    try:
        student_count = Student.objects.count()
        
        return JsonResponse({
            'status': 'healthy',
            'timestamp': datetime.now().isoformat(),
            'database': 'connected',
            'student_count': student_count
        })
    except Exception as e:
        return JsonResponse({
            'status': 'unhealthy',
            'timestamp': datetime.now().isoformat(),
            'error': str(e)
        }, status=500)

# ==================== ADMIN VIEWS ====================

@csrf_exempt
@require_http_methods(["POST"])
def add_student(request):
    """Add a new student WITH DIVISION"""
    try:
        data = json.loads(request.body)
        
        required_fields = ['email', 'first_name', 'last_name', 
                          'password', 'prn_number', 'year_id', 'branch_id', 'semester_id', 'division_id']
        
        for field in required_fields:
            if field not in data or not data[field]:
                return JsonResponse({'error': f'Missing required field: {field}'}, status=400)
        
        if CustomUser.objects.filter(prn_number=data['prn_number']).exists():
            return JsonResponse({'error': 'PRN number already exists'}, status=400)
        
        if Student.objects.filter(prn_number=data['prn_number']).exists():
            return JsonResponse({'error': 'Student with this PRN already exists'}, status=400)
        
        with transaction.atomic():
            username = data.get('username', data['prn_number'])
            
            if CustomUser.objects.filter(username=username).exists():
                username = data['prn_number']
            
            user = CustomUser.objects.create_user(
                username=username,
                email=data['email'],
                first_name=data['first_name'],
                last_name=data['last_name'],
                password=data['password'],
                user_type='student',
                prn_number=data['prn_number']
            )
            
            Student.objects.create(
                user=user,
                prn_number=data['prn_number'],
                year_id=data['year_id'],
                branch_id=data['branch_id'],
                semester_id=data['semester_id'],
                division_id=data['division_id']
            )
        
        return JsonResponse({
            'success': True,
            'message': f'Student added successfully. Can login with PRN: {data["prn_number"]}',
            'username': username,
            'prn_number': data['prn_number']
        }, status=201)
        
    except Exception as e:
        import traceback
        print("ADD STUDENT ERROR:", traceback.format_exc())
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@require_http_methods(["POST"])
def add_teacher(request):
    """Add a new teacher"""
    try:
        data = json.loads(request.body)
        
        required_fields = ['email', 'first_name', 'last_name',
                          'password', 'employee_id']
        
        for field in required_fields:
            if field not in data:
                return JsonResponse({'error': f'Missing required field: {field}'}, status=400)
        
        if CustomUser.objects.filter(prn_number=data['employee_id']).exists():
            return JsonResponse({'error': 'Employee ID already exists'}, status=400)
        
        if Teacher.objects.filter(employee_id=data['employee_id']).exists():
            return JsonResponse({'error': 'Teacher with this Employee ID already exists'}, status=400)
        
        with transaction.atomic():
            username = data.get('username', data['employee_id'])
            
            if CustomUser.objects.filter(username=username).exists():
                username = data['employee_id']
            
            user = CustomUser.objects.create_user(
                username=username,
                email=data['email'],
                first_name=data['first_name'],
                last_name=data['last_name'],
                password=data['password'],
                user_type='teacher',
                prn_number=data['employee_id']
            )
            
            teacher = Teacher.objects.create(
                user=user,
                employee_id=data['employee_id'],
                department_id=data.get('department_id'),
                is_class_teacher=data.get('is_class_teacher', False)
            )
            
            if 'subject_ids' in data and data['subject_ids']:
                teacher.subjects.set(data['subject_ids'])
        
        return JsonResponse({
            'success': True,
            'message': f'Teacher added successfully. Can login with Employee ID: {data["employee_id"]}',
            'username': username,
            'employee_id': data['employee_id'],
            'teacher_id': teacher.id
        }, status=201)
        
    except Exception as e:
        import traceback
        print("ADD TEACHER ERROR:", traceback.format_exc())
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@require_http_methods(["POST"])
def add_subject(request):
    """Add a new subject WITH DIVISION"""
    try:
        data = json.loads(request.body)
        
        required_fields = ['code', 'name', 'year_id', 'branch_id', 'semester_id', 'credits']
        
        for field in required_fields:
            if field not in data or not data[field]:
                return JsonResponse({'error': f'Missing required field: {field}'}, status=400)
        
        semester = get_object_or_404(Semester, id=data['semester_id'], year_id=data['year_id'])
        branch = get_object_or_404(Branch, id=data['branch_id'])
        
        division_id = data.get('division_id') if data.get('division_id') else None
        division = Division.objects.get(id=division_id) if division_id else None
        
        if Subject.objects.filter(
            code=data['code'],
            semester=semester,
            branch=branch,
            division=division
        ).exists():
            return JsonResponse({'error': 'Subject with this code already exists for this branch, semester, and division'}, status=400)
        
        with transaction.atomic():
            subject = Subject.objects.create(
                code=data['code'],
                name=data['name'],
                semester=semester,
                branch=branch,
                division=division,
                credits=int(data['credits'])
            )
        
        div_str = f" for Division {division.name}" if division else " (Common for all divisions)"
        
        return JsonResponse({
            'success': True,
            'message': f'Subject {data["code"]} added successfully{div_str}',
            'subject_id': subject.id
        }, status=201)
        
    except Exception as e:
        import traceback
        print("ADD SUBJECT ERROR:", traceback.format_exc())
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@require_http_methods(["PUT"])
def update_student(request, student_id):
    """Update student information - FIXED"""
    try:
        data = json.loads(request.body)
        student = get_object_or_404(Student, id=student_id)
        
        with transaction.atomic():
            # Update user fields
            if 'first_name' in data:
                student.user.first_name = data['first_name']
            if 'last_name' in data:
                student.user.last_name = data['last_name']
            if 'email' in data:
                student.user.email = data['email']
            student.user.save()
            
            # Update student fields
            if 'year_id' in data:
                student.year_id = data['year_id']
            if 'branch_id' in data:
                student.branch_id = data['branch_id']
            if 'semester_id' in data:
                student.semester_id = data['semester_id']
            if 'division_id' in data:
                student.division_id = data['division_id']
            
            student.save()
            # Force refresh to ensure latest data
            student.refresh_from_db()
        
        return JsonResponse({
            'success': True,
            'message': 'Student updated successfully',
            'updated_student': {
                'id': student.id,
                'prn': student.prn_number,
                'name': student.user.get_full_name(),
                'year': student.year.name,
                'branch': student.branch.name,
                'semester': student.semester.number,
                'division': student.division.name if student.division else 'N/A'
            }
        })
        
    except Exception as e:
        import traceback
        print("UPDATE STUDENT ERROR:", traceback.format_exc())
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@require_http_methods(["PUT"])
def update_teacher(request, teacher_id):
    """Update teacher information"""
    try:
        data = json.loads(request.body)
        teacher = get_object_or_404(Teacher, id=teacher_id)
        
        with transaction.atomic():
            if 'first_name' in data:
                teacher.user.first_name = data['first_name']
            if 'last_name' in data:
                teacher.user.last_name = data['last_name']
            if 'email' in data:
                teacher.user.email = data['email']
            teacher.user.save()
            
            if 'department_id' in data:
                teacher.department_id = data['department_id'] if data['department_id'] else None
            teacher.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Teacher updated successfully'
        })
        
    except Exception as e:
        import traceback
        print("UPDATE TEACHER ERROR:", traceback.format_exc())
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@require_http_methods(["PUT"])
def update_subject(request, subject_id):
    """Update subject information - COMPLETE UPDATE INCLUDING YEAR/BRANCH/SEMESTER"""
    try:
        data = json.loads(request.body)
        subject = get_object_or_404(Subject, id=subject_id)
        
        with transaction.atomic():
            if 'code' in data:
                subject.code = data['code']
            if 'name' in data:
                subject.name = data['name']
            if 'credits' in data:
                subject.credits = int(data['credits'])
            
            # ✅ UPDATE: Handle year and semester changes
            if 'year_id' in data and 'semester_id' in data:
                # Get the semester that belongs to the year
                try:
                    semester = Semester.objects.get(
                        id=data['semester_id'], 
                        year_id=data['year_id']
                    )
                    subject.semester = semester
                except Semester.DoesNotExist:
                    return JsonResponse({
                        'error': 'Invalid semester for the selected year'
                    }, status=400)
            
            # ✅ UPDATE: Handle branch change
            if 'branch_id' in data:
                try:
                    branch = Branch.objects.get(id=data['branch_id'])
                    subject.branch = branch
                except Branch.DoesNotExist:
                    return JsonResponse({
                        'error': 'Invalid branch'
                    }, status=400)
            
            # Handle division
            if 'division_id' in data:
                subject.division_id = data['division_id'] if data['division_id'] else None
            
            subject.save()
        
        # Return updated subject data
        return JsonResponse({
            'success': True,
            'message': 'Subject updated successfully',
            'updated_subject': {
                'id': subject.id,
                'code': subject.code,
                'name': subject.name,
                'year': subject.semester.year.name,
                'branch': subject.branch.name,
                'semester': subject.semester.number,
                'division': subject.division.name if subject.division else 'Common',
                'credits': subject.credits
            }
        })
        
    except Exception as e:
        import traceback
        print("UPDATE SUBJECT ERROR:", traceback.format_exc())
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@require_http_methods(["POST"])
def assign_subjects_to_teacher(request, teacher_id):
    """Assign subjects to a teacher"""
    try:
        data = json.loads(request.body)
        subject_ids = data.get('subject_ids', [])
        
        teacher = get_object_or_404(Teacher, id=teacher_id)
        
        if subject_ids:
            teacher.subjects.set(subject_ids)
        
        return JsonResponse({
            'success': True,
            'message': f'Subjects assigned to {teacher.user.get_full_name()}'
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@require_http_methods(["POST"])
def assign_class_teacher(request):
    """Assign class teacher to students WITH DIVISION"""
    try:
        data = json.loads(request.body)
        teacher_id = data.get('teacher_id')
        year_id = data.get('year_id')
        branch_id = data.get('branch_id')
        semester_id = data.get('semester_id')
        division_id = data.get('division_id')
        
        teacher = get_object_or_404(Teacher, id=teacher_id)
        
        teacher.is_class_teacher = True
        teacher.assigned_class_year_id = year_id
        teacher.assigned_class_branch_id = branch_id
        teacher.assigned_class_semester_id = semester_id
        teacher.assigned_class_division_id = division_id
        teacher.save()
        
        students = Student.objects.filter(
            year_id=year_id,
            branch_id=branch_id,
            semester_id=semester_id,
            division_id=division_id
        )
        
        students.update(class_teacher=teacher)
        
        division = Division.objects.get(id=division_id) if division_id else None
        division_str = f" Division {division.name}" if division else ""
        
        return JsonResponse({
            'success': True,
            'message': f'{teacher.user.get_full_name()} assigned as class teacher to {students.count()} students{division_str}'
        })
        
    except Exception as e:
        import traceback
        print("ASSIGN CLASS TEACHER ERROR:", traceback.format_exc())
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@require_http_methods(["POST"])
def reset_student_password(request, student_id):
    """Reset password for a student"""
    try:
        data = json.loads(request.body)
        new_password = data.get('new_password')
        
        if not new_password or len(new_password) < 6:
            return JsonResponse({'error': 'Password must be at least 6 characters'}, status=400)
        
        student = get_object_or_404(Student, id=student_id)
        student.user.set_password(new_password)
        student.user.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Password reset for {student.prn_number}'
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@require_http_methods(["POST"])
def reset_teacher_password(request, teacher_id):
    """Reset password for a teacher"""
    try:
        data = json.loads(request.body)
        new_password = data.get('new_password')
        
        if not new_password or len(new_password) < 6:
            return JsonResponse({'error': 'Password must be at least 6 characters'}, status=400)
        
        teacher = get_object_or_404(Teacher, id=teacher_id)
        teacher.user.set_password(new_password)
        teacher.user.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Password reset for {teacher.employee_id}'
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

# ==================== DATA APIS ====================

def get_branches(request):
    """Get all branches"""
    try:
        branches = Branch.objects.all().order_by('name')
        return JsonResponse({
            'success': True,
            'branches': [{'id': b.id, 'name': b.name, 'code': b.code} for b in branches]
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def get_years(request):
    """Get all academic years"""
    try:
        years = Year.objects.all().order_by('name')
        return JsonResponse({
            'success': True,
            'years': [{'id': y.id, 'name': y.name} for y in years]
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def get_semesters(request, year_id):
    """Get semesters for a specific year"""
    try:
        semesters = Semester.objects.filter(year_id=year_id).order_by('number')
        return JsonResponse({
            'success': True,
            'semesters': [{'id': s.id, 'number': s.number} for s in semesters]
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def get_divisions(request):
    """Get all divisions"""
    try:
        divisions = Division.objects.all().order_by('name')
        return JsonResponse({
            'success': True,
            'divisions': [{'id': d.id, 'name': d.name} for d in divisions]
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def get_student_subjects(request):
    """Get subjects for student's semester, branch, and division - FIXED"""
    try:
        username = request.GET.get('username')
        
        if not username:
            return JsonResponse({'error': 'Username required'}, status=400)
        
        # Force refresh from database
        user = CustomUser.objects.select_related('student_profile').get(
            username=username, 
            user_type='student'
        )
        student = user.student_profile
        
        # Refresh student object from database to get latest data
        student.refresh_from_db()
        
        # Get subjects for student's division or common subjects (division=None)
        subjects = Subject.objects.filter(
            semester=student.semester,
            branch=student.branch
        ).filter(
            Q(division=student.division) | Q(division__isnull=True)
        ).select_related('semester', 'branch', 'division')
        
        subjects_data = []
        
        for subject in subjects:
            # FIXED: Use TeacherSubject model correctly
            teacher_assignments = TeacherSubject.objects.filter(
                subject=subject
            ).select_related('teacher', 'teacher__user')
            
            if not teacher_assignments.exists():
                subjects_data.append({
                    'id': subject.id,
                    'code': subject.code,
                    'name': subject.name,
                    'credits': subject.credits,
                    'division': subject.division.name if subject.division else 'Common',
                    'teacher': {
                        'id': None,
                        'name': 'No teacher assigned',
                        'employee_id': 'N/A'
                    },
                    'feedback_submitted': False,
                    'no_teacher': True
                })
            else:
                for ts in teacher_assignments:
                    teacher = ts.teacher
                    feedback_exists = Feedback.objects.filter(
                        student=student,
                        subject=subject,
                        teacher=teacher
                    ).exists()
                    
                    subjects_data.append({
                        'id': subject.id,
                        'code': subject.code,
                        'name': subject.name,
                        'credits': subject.credits,
                        'division': subject.division.name if subject.division else 'Common',
                        'teacher': {
                            'id': teacher.id,
                            'name': teacher.user.get_full_name(),
                            'employee_id': teacher.employee_id
                        },
                        'feedback_submitted': feedback_exists,
                        'no_teacher': False
                    })
        
        return JsonResponse({
            'success': True,
            'subjects': subjects_data,
            'student_info': {
                'year': student.year.name,
                'branch': student.branch.name,
                'semester': student.semester.number,
                'division': student.division.name if student.division else 'N/A'
            }
        })
        
    except CustomUser.DoesNotExist:
        return JsonResponse({'error': 'Student not found'}, status=404)
    except Exception as e:
        import traceback
        print("GET STUDENT SUBJECTS ERROR:", traceback.format_exc())
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@require_http_methods(["POST"])
def submit_feedback(request):
    """Submit student feedback with sentiment analysis"""
    try:
        data = json.loads(request.body)
        username = data.get('username')
        
        if not username:
            return JsonResponse({'error': 'Username is required'}, status=400)
        
        user = CustomUser.objects.get(username=username, user_type='student')
        student = user.student_profile
        
        subject = get_object_or_404(Subject, id=data['subject_id'])
        teacher = get_object_or_404(Teacher, id=data['teacher_id'])
        
        if Feedback.objects.filter(
            student=student,
            subject=subject,
            teacher=teacher,
            semester=student.semester
        ).exists():
            return JsonResponse({
                'error': 'You have already submitted feedback for this subject and teacher'
            }, status=400)
        
        comments = data.get('comments', '').strip()
        suggestions = data.get('suggestions', '').strip()
        
        comment_sentiment, comment_score = analyze_sentiment(comments)
        suggestion_sentiment, suggestion_score = analyze_sentiment(suggestions)
        
        feedback = Feedback.objects.create(
            student=student,
            teacher=teacher,
            subject=subject,
            semester=student.semester,
            teaching_effectiveness=int(data['teaching_effectiveness']),
            course_content=int(data['course_content']),
            interaction_quality=int(data['interaction_quality']),
            assignment_feedback=int(data['assignment_feedback']),
            overall_satisfaction=int(data['overall_satisfaction']),
            comments=comments,
            comment_sentiment=comment_sentiment,
            comment_sentiment_score=comment_score,
            suggestions=suggestions,
            suggestion_sentiment=suggestion_sentiment,
            suggestion_sentiment_score=suggestion_score,
            is_anonymous=data.get('is_anonymous', True)
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Feedback submitted successfully',
            'feedback_id': feedback.id,
            'sentiment': {
                'comment': comment_sentiment,
                'suggestion': suggestion_sentiment
            }
        }, status=201)
        
    except KeyError as e:
        return JsonResponse({'error': f'Missing required field: {str(e)}'}, status=400)
    except Exception as e:
        import traceback
        print("SUBMIT FEEDBACK ERROR:", traceback.format_exc())
        return JsonResponse({'error': str(e)}, status=500)

# ==================== TEACHER VIEWS ====================

def teacher_dashboard(request):
    """Get teacher dashboard with analytics"""
    try:
        username = request.GET.get('username')
        
        if not username:
            return JsonResponse({'error': 'Username required'}, status=400)
        
        user = CustomUser.objects.get(username=username, user_type='teacher')
        teacher = user.teacher_profile
        
        feedbacks = Feedback.objects.filter(teacher=teacher)
        total_feedback = feedbacks.count()
        
        if total_feedback > 0:
            avg_ratings = feedbacks.aggregate(
                avg_teaching=Avg('teaching_effectiveness'),
                avg_content=Avg('course_content'),
                avg_interaction=Avg('interaction_quality'),
                avg_assignment=Avg('assignment_feedback'),
                avg_overall=Avg('overall_satisfaction')
            )
            
            positive_count = 0
            negative_count = 0
            neutral_count = 0
            
            for fb in feedbacks:
                if fb.comment_sentiment == 'positive':
                    positive_count += 1
                elif fb.comment_sentiment == 'negative':
                    negative_count += 1
                elif fb.comment_sentiment == 'neutral':
                    neutral_count += 1
                
                if fb.suggestion_sentiment == 'positive':
                    positive_count += 1
                elif fb.suggestion_sentiment == 'negative':
                    negative_count += 1
                elif fb.suggestion_sentiment == 'neutral':
                    neutral_count += 1
            
            sentiment_stats = {
                'positive': positive_count,
                'negative': negative_count,
                'neutral': neutral_count,
            }
            
            rating_distribution = []
            for rating in range(1, 6):
                count = feedbacks.filter(overall_satisfaction=rating).count()
                rating_distribution.append({'rating': rating, 'count': count})
            
        else:
            avg_ratings = {
                'avg_teaching': 0, 'avg_content': 0, 'avg_interaction': 0,
                'avg_assignment': 0, 'avg_overall': 0
            }
            sentiment_stats = {'positive': 0, 'negative': 0, 'neutral': 0}
            rating_distribution = [{'rating': i, 'count': 0} for i in range(1, 6)]
        
        subjects_taught = teacher.subjects.count()
        
        return JsonResponse({
            'success': True,
            'teacher': {
                'employee_id': teacher.employee_id,
                'name': user.get_full_name(),
                'department': teacher.department.name if teacher.department else 'N/A',
                'email': user.email,
                'is_class_teacher': teacher.is_class_teacher
            },
            'statistics': {
                'total_feedback': total_feedback,
                'subjects_taught': subjects_taught,
                'average_ratings': {
                    'teaching_effectiveness': round(avg_ratings['avg_teaching'] or 0, 2),
                    'course_content': round(avg_ratings['avg_content'] or 0, 2),
                    'interaction_quality': round(avg_ratings['avg_interaction'] or 0, 2),
                    'assignment_feedback': round(avg_ratings['avg_assignment'] or 0, 2),
                    'overall_satisfaction': round(avg_ratings['avg_overall'] or 0, 2)
                },
                'sentiment_distribution': sentiment_stats,
                'rating_distribution': rating_distribution
            }
        })
        
    except CustomUser.DoesNotExist:
        return JsonResponse({'error': 'Teacher not found'}, status=404)
    except Exception as e:
        import traceback
        print("ERROR:", traceback.format_exc())
        return JsonResponse({'error': str(e)}, status=500)

def teacher_feedback_data(request):
    """Get detailed feedback data for teacher"""
    try:
        username = request.GET.get('username')
        
        if not username:
            return JsonResponse({'error': 'Username required'}, status=400)
        
        user = CustomUser.objects.get(username=username, user_type='teacher')
        teacher = user.teacher_profile
        
        feedbacks = Feedback.objects.filter(teacher=teacher).select_related(
            'subject', 'semester', 'semester__year', 'student', 'student__user',
            'student__branch', 'student__year', 'student__division'
        ).order_by('-created_at')
        
        for fb in feedbacks:
            if fb.comments and not fb.comment_sentiment:
                fb.comment_sentiment, fb.comment_sentiment_score = analyze_sentiment(fb.comments)
                fb.save()
            if fb.suggestions and not fb.suggestion_sentiment:
                fb.suggestion_sentiment, fb.suggestion_sentiment_score = analyze_sentiment(fb.suggestions)
                fb.save()
        
        feedback_data = []
        for fb in feedbacks:
            feedback_data.append({
                'id': fb.id,
                'student_prn': fb.student.prn_number if not fb.is_anonymous else 'Anonymous',
                'student_name': fb.student.user.get_full_name() if not fb.is_anonymous else 'Anonymous',
                'student_branch': fb.student.branch.name if not fb.is_anonymous else 'N/A',
                'student_year': fb.student.year.name if not fb.is_anonymous else 'N/A',
                'student_division': fb.student.division.name if (not fb.is_anonymous and fb.student.division) else 'N/A',
                'subject': {
                    'code': fb.subject.code,
                    'name': fb.subject.name
                },
                'semester': f"{fb.semester.year.name} - Semester {fb.semester.number}",
                'ratings': {
                    'teaching_effectiveness': fb.teaching_effectiveness,
                    'course_content': fb.course_content,
                    'interaction_quality': fb.interaction_quality,
                    'assignment_feedback': fb.assignment_feedback,
                    'overall_satisfaction': fb.overall_satisfaction
                },
                'comments': fb.comments or '',
                'comment_sentiment': fb.comment_sentiment,
                'comment_sentiment_score': round(fb.comment_sentiment_score, 3) if fb.comment_sentiment_score else 0,
                'suggestions': fb.suggestions or '',
                'suggestion_sentiment': fb.suggestion_sentiment,
                'suggestion_sentiment_score': round(fb.suggestion_sentiment_score, 3) if fb.suggestion_sentiment_score else 0,
                'is_anonymous': fb.is_anonymous,
                'created_at': fb.created_at.strftime('%Y-%m-%d %H:%M:%S')
            })
        
        total_feedback = feedbacks.count()
        
        positive_count = 0
        negative_count = 0
        neutral_count = 0
        
        for fb in feedbacks:
            if fb.comment_sentiment == 'positive':
                positive_count += 1
            elif fb.comment_sentiment == 'negative':
                negative_count += 1
            elif fb.comment_sentiment == 'neutral':
                neutral_count += 1
            
            if fb.suggestion_sentiment == 'positive':
                positive_count += 1
            elif fb.suggestion_sentiment == 'negative':
                negative_count += 1
            elif fb.suggestion_sentiment == 'neutral':
                neutral_count += 1
        
        sentiment_stats = {
            'positive': positive_count,
            'negative': negative_count,
            'neutral': neutral_count,
        }
        
        rating_distribution = []
        for rating in range(1, 6):
            count = feedbacks.filter(overall_satisfaction=rating).count()
            rating_distribution.append({'rating': rating, 'count': count})
        
        subject_stats = feedbacks.values(
            'subject__code', 'subject__name'
        ).annotate(
            feedback_count=Count('id'),
            avg_rating=Avg('overall_satisfaction')
        ).order_by('-feedback_count')
        
        return JsonResponse({
            'success': True,
            'feedback': feedback_data,
            'statistics': {
                'total_feedback': total_feedback,
                'sentiment_stats': sentiment_stats,
                'rating_distribution': rating_distribution,
                'subject_breakdown': list(subject_stats)
            },
            'sentiment_stats': sentiment_stats
        })
        
    except CustomUser.DoesNotExist:
        return JsonResponse({'error': 'Teacher not found'}, status=404)
    except Exception as e:
        import traceback
        print("ERROR:", traceback.format_exc())
        return JsonResponse({'error': str(e)}, status=500)

def download_feedback_data(request):
    """Download feedback data as CSV"""
    try:
        username = request.GET.get('username')
        
        if not username:
            return JsonResponse({'error': 'Username required'}, status=400)
        
        user = CustomUser.objects.get(username=username, user_type='teacher')
        teacher = user.teacher_profile
        
        feedbacks = Feedback.objects.filter(teacher=teacher).select_related(
            'student', 'student__user', 'student__branch', 'student__year', 'student__division',
            'subject', 'semester', 'semester__year'
        ).order_by('-created_at')
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="feedback_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        response.write('\ufeff')
        
        writer = csv.writer(response)
        writer.writerow([
            'ID', 'Date', 
             #'PRN',
             #'Student Name',
              'Year', 'Branch', 'Division', 'Subject Code',
            'Subject Name', 'Semester', 'Teaching', 'Content', 'Interaction',
            'Assignment', 'Overall', 'Comments', 'Comment Sentiment', 'Comment Score',
            'Suggestions', 'Suggestion Sentiment', 'Suggestion Score', 'Is Anonymous'
        ])
        
        for fb in feedbacks:
            writer.writerow([
                fb.id,
                fb.created_at.strftime('%d-%m-%Y %H:%M'),
                #fb.student.prn_number,
                #fb.student.user.get_full_name(),
                fb.student.year.name,
                fb.student.branch.name,
                fb.student.division.name if fb.student.division else 'N/A',
                fb.subject.code,
                fb.subject.name,
                f"{fb.semester.year.name} Sem-{fb.semester.number}",
                fb.teaching_effectiveness,
                fb.course_content,
                fb.interaction_quality,
                fb.assignment_feedback,
                fb.overall_satisfaction,
                fb.comments or 'No comments',
                fb.comment_sentiment or 'Not analyzed',
                f"{fb.comment_sentiment_score:.2f}" if fb.comment_sentiment_score else '0.00',
                fb.suggestions or 'No suggestions',
                fb.suggestion_sentiment or 'Not analyzed',
                f"{fb.suggestion_sentiment_score:.2f}" if fb.suggestion_sentiment_score else '0.00',
                'Yes' if fb.is_anonymous else 'No'
            ])
        
        return response
        
    except CustomUser.DoesNotExist:
        return JsonResponse({'error': 'Teacher not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

# ==================== CLASS TEACHER VIEWS ====================

def class_teacher_dashboard(request):
    """Get class teacher specific dashboard data WITH DIVISION"""
    try:
        username = request.GET.get('username')
        
        if not username:
            return JsonResponse({'error': 'Username required'}, status=400)
        
        user = CustomUser.objects.get(username=username, user_type='teacher')
        teacher = user.teacher_profile
        
        if not teacher.is_class_teacher:
            return JsonResponse({'error': 'Not authorized as class teacher'}, status=403)
        
        # Get assigned class students
        students = Student.objects.filter(
            class_teacher=teacher
        ).select_related('user', 'year', 'branch', 'semester', 'division')
        
        # If no class_teacher assignment, use year/branch/semester/division
        if not students.exists() and teacher.assigned_class_year:
            students = Student.objects.filter(
                year=teacher.assigned_class_year,
                branch=teacher.assigned_class_branch,
                semester=teacher.assigned_class_semester,
                division=teacher.assigned_class_division
            ).select_related('user', 'year', 'branch', 'semester', 'division')
        
        total_students = students.count()
        
        # Get all subjects for this class
        if total_students > 0:
            first_student = students.first()
            subjects = Subject.objects.filter(
                branch=first_student.branch,
                semester=first_student.semester
            ).filter(
                Q(division=first_student.division) | Q(division__isnull=True)
            ).prefetch_related('teachers')
        else:
            subjects = Subject.objects.none()
        
        # Calculate feedback statistics
        students_submitted = set()
        students_pending = []
        
        for student in students:
            student_feedbacks = Feedback.objects.filter(student=student)
            
            if student_feedbacks.exists():
                students_submitted.add(student.id)
            else:
                students_pending.append({
                    'prn': student.prn_number,
                    'name': student.user.get_full_name(),
                    'email': student.user.email
                })
        
        submitted_count = len(students_submitted)
        pending_count = total_students - submitted_count
        completion_rate = (submitted_count / total_students * 100) if total_students > 0 else 0
        
        division_str = teacher.assigned_class_division.name if teacher.assigned_class_division else 'N/A'
        
        return JsonResponse({
            'success': True,
            'class_info': {
                'year': teacher.assigned_class_year.name if teacher.assigned_class_year else 'N/A',
                'branch': teacher.assigned_class_branch.name if teacher.assigned_class_branch else 'N/A',
                'semester': f"Semester {teacher.assigned_class_semester.number}" if teacher.assigned_class_semester else 'N/A',
                'division': division_str,
            },
            'statistics': {
                'total_students': total_students,
                'submitted_feedback': submitted_count,
                'pending_feedback': pending_count,
                'completion_rate': round(completion_rate, 2),
                'total_subjects': subjects.count()
            },
            'pending_students': students_pending[:10]
        })
        
    except CustomUser.DoesNotExist:
        return JsonResponse({'error': 'Teacher not found'}, status=404)
    except Exception as e:
        import traceback
        print("ERROR:", traceback.format_exc())
        return JsonResponse({'error': str(e)}, status=500)

def class_teacher_student_tracking(request):
    """Get detailed student tracking for class teacher WITH DIVISION"""
    try:
        username = request.GET.get('username')
        
        if not username:
            return JsonResponse({'error': 'Username required'}, status=400)
        
        user = CustomUser.objects.get(username=username, user_type='teacher')
        teacher = user.teacher_profile
        
        if not teacher.is_class_teacher:
            return JsonResponse({'error': 'Not authorized as class teacher'}, status=403)
        
        # Get assigned class students
        students = Student.objects.filter(
            class_teacher=teacher
        ).select_related('user', 'year', 'branch', 'semester', 'division')
        
        if not students.exists() and teacher.assigned_class_year:
            students = Student.objects.filter(
                year=teacher.assigned_class_year,
                branch=teacher.assigned_class_branch,
                semester=teacher.assigned_class_semester,
                division=teacher.assigned_class_division
            ).select_related('user', 'year', 'branch', 'semester', 'division')
        
        # Get all subjects for class
        if students.exists():
            first_student = students.first()
            subjects = Subject.objects.filter(
                branch=first_student.branch,
                semester=first_student.semester
            ).filter(
                Q(division=first_student.division) | Q(division__isnull=True)
            ).prefetch_related('teachers')
        else:
            subjects = Subject.objects.none()
        
        total_possible_feedbacks = subjects.count() * Teacher.objects.filter(subjects__in=subjects).distinct().count()
        
        students_data = []
        
        for student in students:
            feedbacks_given = Feedback.objects.filter(student=student)
            feedback_count = feedbacks_given.count()
            
            # Calculate which subjects have feedback
            subjects_with_feedback = []
            subjects_without_feedback = []
            
            for subject in subjects:
                teachers_for_subject = Teacher.objects.filter(subjects=subject)
                for teacher_obj in teachers_for_subject:
                    has_feedback = Feedback.objects.filter(
                        student=student,
                        subject=subject,
                        teacher=teacher_obj
                    ).exists()
                    
                    if has_feedback:
                        subjects_with_feedback.append(f"{subject.code} ({teacher_obj.employee_id})")
                    else:
                        subjects_without_feedback.append(f"{subject.code} ({teacher_obj.employee_id})")
            
            students_data.append({
                'prn': student.prn_number,
                'name': student.user.get_full_name(),
                'email': student.user.email,
                'division': student.division.name if student.division else 'N/A',
                'feedback_submitted': feedback_count,
                'feedback_pending': len(subjects_without_feedback),
                'completion_percentage': round((feedback_count / total_possible_feedbacks * 100), 2) if total_possible_feedbacks > 0 else 0,
                'subjects_completed': subjects_with_feedback,
                'subjects_pending': subjects_without_feedback,
                'status': 'Complete' if len(subjects_without_feedback) == 0 else 'Pending'
            })
        
        return JsonResponse({
            'success': True,
            'students': students_data,
            'summary': {
                'total_students': len(students_data),
                'total_subjects': subjects.count(),
                'completed_students': len([s for s in students_data if s['status'] == 'Complete']),
                'pending_students': len([s for s in students_data if s['status'] == 'Pending'])
            }
        })
        
    except CustomUser.DoesNotExist:
        return JsonResponse({'error': 'Teacher not found'}, status=404)
    except Exception as e:
        import traceback
        print("ERROR:", traceback.format_exc())
        return JsonResponse({'error': str(e)}, status=500)

def download_class_teacher_report(request):
    """Download Excel report for class teacher - ENHANCED WITH ROLL NO & STUDENT NAMES"""
    try:
        username = request.GET.get('username')
        
        if not username:
            return JsonResponse({'error': 'Username required'}, status=400)
        
        user = CustomUser.objects.get(username=username, user_type='teacher')
        teacher = user.teacher_profile
        
        if not teacher.is_class_teacher:
            return JsonResponse({'error': 'Not authorized as class teacher'}, status=403)
        
        # Get assigned class students
        students = Student.objects.filter(
            class_teacher=teacher
        ).select_related('user', 'year', 'branch', 'semester', 'division')
        
        if not students.exists() and teacher.assigned_class_year:
            students = Student.objects.filter(
                year=teacher.assigned_class_year,
                branch=teacher.assigned_class_branch,
                semester=teacher.assigned_class_semester,
                division=teacher.assigned_class_division
            ).select_related('user', 'year', 'branch', 'semester', 'division')
        
        # Get subjects
        if students.exists():
            first_student = students.first()
            subjects = Subject.objects.filter(
                branch=first_student.branch,
                semester=first_student.semester
            ).filter(
                Q(division=first_student.division) | Q(division__isnull=True)
            ).prefetch_related('teachers')
        else:
            subjects = Subject.objects.none()
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        
        # Sheet 1: Summary
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        # Summary info
        ws_summary['A1'] = 'CLASS FEEDBACK TRACKING REPORT'
        ws_summary['A1'].font = Font(bold=True, size=16)
        ws_summary.merge_cells('A1:D1')
        
        ws_summary['A3'] = 'Class Teacher:'
        ws_summary['B3'] = teacher.user.get_full_name()
        ws_summary['A4'] = 'Employee ID:'
        ws_summary['B4'] = teacher.employee_id
        ws_summary['A5'] = 'Class:'
        division_name = teacher.assigned_class_division.name if teacher.assigned_class_division else 'N/A'
        ws_summary['B5'] = f"{teacher.assigned_class_year.name if teacher.assigned_class_year else 'N/A'} - {teacher.assigned_class_branch.name if teacher.assigned_class_branch else 'N/A'} - Semester {teacher.assigned_class_semester.number if teacher.assigned_class_semester else 'N/A'} - Division {division_name}"
        ws_summary['A6'] = 'Report Date:'
        ws_summary['B6'] = datetime.now().strftime('%d-%m-%Y %H:%M')
        
        # Statistics
        total_students = students.count()
        students_with_feedback = set()
        total_feedbacks = 0
        
        for student in students:
            feedbacks = Feedback.objects.filter(student=student)
            if feedbacks.exists():
                students_with_feedback.add(student.id)
                total_feedbacks += feedbacks.count()
        
        ws_summary['A8'] = 'STATISTICS'
        ws_summary['A8'].font = Font(bold=True, size=14)
        
        ws_summary['A10'] = 'Total Students:'
        ws_summary['B10'] = total_students
        ws_summary['A11'] = 'Students Submitted Feedback:'
        ws_summary['B11'] = len(students_with_feedback)
        ws_summary['A12'] = 'Students Pending:'
        ws_summary['B12'] = total_students - len(students_with_feedback)
        ws_summary['A13'] = 'Total Feedbacks Received:'
        ws_summary['B13'] = total_feedbacks
        ws_summary['A14'] = 'Completion Rate:'
        ws_summary['B14'] = f"{round((len(students_with_feedback) / total_students * 100), 2)}%" if total_students > 0 else "0%"
        
        # Sheet 2: Students Submitted Feedback (ENHANCED)
        ws_submitted = wb.create_sheet("Students Submitted")
        
        headers_submitted = ['Sr.No', 'Roll No', 'PRN', 'Student Name', 'Email', 'Year', 'Branch', 'Semester', 'Division', 'Total Feedbacks', 'Status']
        for col, header in enumerate(headers_submitted, 1):
            cell = ws_submitted.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        row = 2
        sr_no = 1
        # Sort students by roll number if available, otherwise by PRN
        sorted_students = sorted(students, key=lambda s: (s.prn_number))
        
        for student in sorted_students:
            feedbacks = Feedback.objects.filter(student=student)
            if feedbacks.exists():
                ws_submitted.cell(row=row, column=1, value=sr_no)
                # Roll Number (using PRN as roll number - modify if you have separate field)
                ws_submitted.cell(row=row, column=2, value=student.prn_number)
                ws_submitted.cell(row=row, column=3, value=student.prn_number)
                ws_submitted.cell(row=row, column=4, value=student.user.get_full_name())
                ws_submitted.cell(row=row, column=5, value=student.user.email)
                ws_submitted.cell(row=row, column=6, value=student.year.name)
                ws_submitted.cell(row=row, column=7, value=student.branch.name)
                ws_submitted.cell(row=row, column=8, value=f"Semester {student.semester.number}")
                ws_submitted.cell(row=row, column=9, value=student.division.name if student.division else 'N/A')
                ws_submitted.cell(row=row, column=10, value=feedbacks.count())
                
                status_cell = ws_submitted.cell(row=row, column=11, value='Submitted')
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                status_cell.font = Font(color="006100", bold=True)
                
                row += 1
                sr_no += 1
        
        # Sheet 3: Students Pending Feedback (ENHANCED)
        ws_pending = wb.create_sheet("Students Pending")
        
        headers_pending = ['Sr.No', 'Roll No', 'PRN', 'Student Name', 'Email', 'Year', 'Branch', 'Semester', 'Division', 'Status']
        for col, header in enumerate(headers_pending, 1):
            cell = ws_pending.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        row = 2
        sr_no = 1
        for student in sorted_students:
            feedbacks = Feedback.objects.filter(student=student)
            if not feedbacks.exists():
                ws_pending.cell(row=row, column=1, value=sr_no)
                # Roll Number
                ws_pending.cell(row=row, column=2, value=student.prn_number)
                ws_pending.cell(row=row, column=3, value=student.prn_number)
                ws_pending.cell(row=row, column=4, value=student.user.get_full_name())
                ws_pending.cell(row=row, column=5, value=student.user.email)
                ws_pending.cell(row=row, column=6, value=student.year.name)
                ws_pending.cell(row=row, column=7, value=student.branch.name)
                ws_pending.cell(row=row, column=8, value=f"Semester {student.semester.number}")
                ws_pending.cell(row=row, column=9, value=student.division.name if student.division else 'N/A')
                
                status_cell = ws_pending.cell(row=row, column=10, value='Pending')
                status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                status_cell.font = Font(color="9C0006", bold=True)
                
                row += 1
                sr_no += 1
        
        # Sheet 4: Detailed Student Tracking (ENHANCED)
        ws_detailed = wb.create_sheet("Detailed Tracking")
        
        headers_detailed = ['Roll No', 'PRN', 'Student Name', 'Email', 'Division']
        
        # Add subject columns
        subject_teacher_combos = []
        for subject in subjects:
            teachers_for_subject = Teacher.objects.filter(subjects=subject)
            for teacher_obj in teachers_for_subject:
                combo_name = f"{subject.code}\n({teacher_obj.employee_id})"
                headers_detailed.append(combo_name)
                subject_teacher_combos.append((subject, teacher_obj))
        
        headers_detailed.extend(['Total Submitted', 'Total Pending', 'Completion %'])
        
        # Write headers
        for col, header in enumerate(headers_detailed, 1):
            cell = ws_detailed.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Write student data
        row = 2
        for student in sorted_students:
            ws_detailed.cell(row=row, column=1, value=student.prn_number)  # Roll No
            ws_detailed.cell(row=row, column=2, value=student.prn_number)  # PRN
            ws_detailed.cell(row=row, column=3, value=student.user.get_full_name())
            ws_detailed.cell(row=row, column=4, value=student.user.email)
            ws_detailed.cell(row=row, column=5, value=student.division.name if student.division else 'N/A')
            
            col = 6
            submitted_count = 0
            pending_count = 0
            
            for subject, teacher_obj in subject_teacher_combos:
                has_feedback = Feedback.objects.filter(
                    student=student,
                    subject=subject,
                    teacher=teacher_obj
                ).exists()
                
                cell = ws_detailed.cell(row=row, column=col)
                if has_feedback:
                    cell.value = '✓'
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    cell.font = Font(color="006100", bold=True, size=14)
                    submitted_count += 1
                else:
                    cell.value = '✗'
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    cell.font = Font(color="9C0006", bold=True, size=14)
                    pending_count += 1
                
                cell.alignment = Alignment(horizontal='center', vertical='center')
                col += 1
            
            ws_detailed.cell(row=row, column=col, value=submitted_count)
            ws_detailed.cell(row=row, column=col+1, value=pending_count)
            
            completion_percent = (submitted_count / (submitted_count + pending_count) * 100) if (submitted_count + pending_count) > 0 else 0
            ws_detailed.cell(row=row, column=col+2, value=f"{round(completion_percent, 2)}%")
            
            row += 1
        
        # Sheet 5: Complete Student List (NEW)
        ws_complete_list = wb.create_sheet("Complete Student List")
        
        headers_complete = ['Sr.No', 'Roll No', 'PRN', 'Student Name', 'Email', 'Year', 'Branch', 'Semester', 'Division', 'Feedbacks Given', 'Status']
        for col, header in enumerate(headers_complete, 1):
            cell = ws_complete_list.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        row = 2
        for idx, student in enumerate(sorted_students, 1):
            feedbacks = Feedback.objects.filter(student=student)
            has_feedback = feedbacks.exists()
            
            ws_complete_list.cell(row=row, column=1, value=idx)
            ws_complete_list.cell(row=row, column=2, value=student.prn_number)
            ws_complete_list.cell(row=row, column=3, value=student.prn_number)
            ws_complete_list.cell(row=row, column=4, value=student.user.get_full_name())
            ws_complete_list.cell(row=row, column=5, value=student.user.email)
            ws_complete_list.cell(row=row, column=6, value=student.year.name)
            ws_complete_list.cell(row=row, column=7, value=student.branch.name)
            ws_complete_list.cell(row=row, column=8, value=f"Semester {student.semester.number}")
            ws_complete_list.cell(row=row, column=9, value=student.division.name if student.division else 'N/A')
            ws_complete_list.cell(row=row, column=10, value=feedbacks.count() if has_feedback else 0)
            
            status_cell = ws_complete_list.cell(row=row, column=11, value='Submitted' if has_feedback else 'Pending')
            if has_feedback:
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                status_cell.font = Font(color="006100", bold=True)
            else:
                status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                status_cell.font = Font(color="9C0006", bold=True)
            
            row += 1
        
        # Adjust column widths for all sheets
        for ws in [ws_summary, ws_submitted, ws_pending, ws_detailed, ws_complete_list]:
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="Class_Feedback_Report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        wb.save(response)
        return response
        
    except CustomUser.DoesNotExist:
        return JsonResponse({'error': 'Teacher not found'}, status=404)
    except Exception as e:
        import traceback
        print("ERROR:", traceback.format_exc())
        return JsonResponse({'error': str(e)}, status=500)